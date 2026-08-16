#!/usr/bin/env python3
import importlib.util
import tempfile
import unittest
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
spec = importlib.util.spec_from_file_location("unified_plan", ROOT / "scripts" / "unified_plan.py")
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)


class UnifiedPlanTest(unittest.TestCase):
    def test_weekdays_between_count_person_days(self):
        weeks = mod.weeks_from_range(date(2026, 8, 3), date(2026, 8, 14))
        self.assertEqual(weeks["2026-08-03"], "5")
        self.assertEqual(weeks["2026-08-10"], "5")

    def test_person_days_cap_uses_first_weekdays(self):
        weeks = mod.weeks_from_range(date(2026, 8, 3), date(2026, 8, 14), person_days=7)
        self.assertEqual(weeks["2026-08-03"], "5")
        self.assertEqual(weeks["2026-08-10"], "2")

    def test_vacation_range_marks_otpusk(self):
        weeks = mod.weeks_from_range(date(2026, 8, 10), date(2026, 8, 14), vacation=True)
        self.assertEqual(weeks["2026-08-10"], "отпуск")

    def test_type_and_role_lists_drop_cko_and_vacation(self):
        self.assertEqual(mod.TYPES, ["деливери", "дискавери"])
        self.assertEqual(mod.ROLES, ["Дизайн", "SA", "PO", "BE", "FE", "QA", "Контент"])
        self.assertEqual(mod.canonical_type("ЦКО"), "")
        self.assertEqual(mod.canonical_type("деливери"), "деливери")
        self.assertEqual(mod.canonical_role("Отпуск"), "")
        self.assertTrue(mod.parse_vacation_flag("да"))
        self.assertFalse(mod.parse_vacation_flag(""))

    def test_parse_canonical_team_aliases(self):
        self.assertEqual(mod.canonical_team("Фикса"), "ДИ")
        self.assertEqual(mod.canonical_team("Репрайсы"), "Монетизация")
        self.assertEqual(mod.canonical_team("Домашний интернет"), "ДИ")
        self.assertEqual(mod.canonical_team("Тарифы"), "Тарифы")

    def test_round_trip_excel_keeps_roles_names_and_week_days(self):
        entries = [
            {
                "team": "Тарифы",
                "task": "Карточка тарифа",
                "resource": "Шлотгауэр Иван",
                "role": "QA",
                "ticket": "B2CPROD-1",
                "type": "деливери",
                "weeks": {"2026-08-31": "5", "2026-09-07": "2"},
            },
            {
                "team": "ДИ",
                "task": "FMC",
                "resource": "Егор",
                "role": "Дизайн",
                "ticket": "",
                "type": "",
                "vacation": True,
                "weeks": {"2026-08-03": "отпуск"},
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "plan.xlsx"
            mod.write_workbook(path, entries=entries, include_examples=False)
            parsed = mod.parse_workbook(path)
        by_task = {item["task"]: item for item in parsed}
        qa = by_task["Карточка тарифа"]
        self.assertEqual(qa["team"], "Тарифы")
        self.assertEqual(qa["resource"], "Шлотгауэр Иван")
        self.assertEqual(qa["role"], "QA")
        self.assertEqual(qa["weeks"]["2026-08-31"], "5")
        self.assertEqual(qa["weeks"]["2026-09-07"], "2")
        self.assertEqual(qa["type"], "деливери")
        self.assertFalse(qa.get("vacation"))
        vac = by_task["FMC"]
        self.assertEqual(vac["team"], "ДИ")
        self.assertEqual(vac["role"], "Дизайн")
        self.assertTrue(vac["vacation"])
        self.assertEqual(vac["weeks"]["2026-08-03"], "отпуск")

    def test_start_end_row_without_week_cells(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "План"
        ws.append(mod.PLAN_HEADERS)
        ws.append(
            [
                "Монетизация",
                "Орига",
                "",
                "деливери",
                "Роган Татьяна",
                "SA",
                "",
                date(2026, 8, 3),
                date(2026, 8, 7),
                3,
                "",
            ]
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "dates.xlsx"
            wb.save(path)
            parsed = mod.parse_workbook(path)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["team"], "Монетизация")
        self.assertEqual(parsed[0]["resource"], "Роган Татьяна")
        self.assertEqual(parsed[0]["role"], "SA")
        self.assertEqual(parsed[0]["weeks"]["2026-08-03"], "3")

    def test_example_rows_are_skipped(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "План"
        ws.append(mod.PLAN_HEADERS)
        ws.append(
            ["Тарифы", "ПРИМЕР: не грузить", "", "", "Косенко Данил", "FE", "", date(2026, 8, 3), date(2026, 8, 7), "", ""]
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "example.xlsx"
            wb.save(path)
            self.assertEqual(mod.parse_workbook(path), [])

    def test_aliases_unify_team_role_and_person(self):
        self.assertEqual(mod.canonical_team("Фикса"), "ДИ")
        self.assertEqual(mod.canonical_role("аналитика"), "SA")
        self.assertEqual(mod.canonical_role("фронт"), "FE")
        self.assertEqual(mod.canonical_resource("Шлогауэр"), "Шлотгауэр Иван")
        self.assertEqual(mod.canonical_resource("Касенко"), "Косенко Данил")
        self.assertEqual(mod.canonical_resource("Потребность test ЦКО"), "Потребность QA")
        self.assertEqual(mod.canonical_resource("Шлотгауэр Иван Александрович"), "Шлотгауэр Иван")
        self.assertEqual(mod.role_for_resource("Шлогауэр"), "QA")
        self.assertEqual(mod.canonical_resource("Роган Татьяна Михайловна"), "Роган Татьяна")
        self.assertEqual(mod.canonical_resource("Володя (SA) нужен еще один"), "Володя (SA) нужен еще один")
        self.assertEqual(mod.canonical_resource("в помощь Володе"), "в помощь Володе")

    def test_workbook_has_instruction_plan_and_reference(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "tpl.xlsx"
            mod.write_workbook(path, include_examples=True)
            from openpyxl import load_workbook

            wb = load_workbook(path)
            self.assertEqual(wb.sheetnames, ["Инструкция", "Справочник", "План"])
            plan = wb["План"]
            self.assertEqual(plan["A1"].value, "Команда")
            self.assertEqual(plan["G1"].value, "Отпуск")
            self.assertTrue(str(plan["L1"].value).count("."))
            self.assertTrue(str(plan["B2"].value).startswith("ПРИМЕР"))
            example_teams = {
                plan.cell(r, 1).value
                for r in range(2, plan.max_row + 1)
                if str(plan.cell(r, 2).value or "").startswith("ПРИМЕР")
            }
            self.assertEqual(example_teams, {"МегаИнтернет"})
            example_tasks = {
                plan.cell(r, 2).value
                for r in range(2, plan.max_row + 1)
                if str(plan.cell(r, 2).value or "").startswith("ПРИМЕР")
            }
            self.assertIn("ПРИМЕР: 5G кино", example_tasks)
            self.assertGreaterEqual(len(example_tasks), 5)
            self.assertEqual(mod.parse_workbook(path), [])
            types = [row[0] for row in wb["Справочник"].iter_rows(min_row=2, min_col=3, max_col=3, values_only=True) if row[0]]
            roles = [row[0] for row in wb["Справочник"].iter_rows(min_row=2, min_col=7, max_col=7, values_only=True) if row[0]]
            self.assertEqual(types, ["деливери", "дискавери"])
            self.assertEqual(roles, ["Дизайн", "SA", "PO", "BE", "FE", "QA", "Контент"])
            people = [row[0] for row in wb["Справочник"].iter_rows(min_row=2, min_col=5, max_col=5, values_only=True) if row[0]]
            self.assertIn("Шлотгауэр Иван", people)
            self.assertIn("Косенко Данил", people)
            self.assertIn("Титов Иван", people)
            self.assertIn("Потребность QA", people)
            week_lists = [dv.formula1 for dv in plan.data_validations.dataValidation]
            self.assertTrue(any(item == '"1,2,3,4,5"' for item in week_lists))
            self.assertFalse(any("отпуск" in (item or "").lower() and "1,2,3,4,5" in (item or "") for item in week_lists))

    def test_parse_incoming_skips_old_color_workbooks(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            incoming = Path(tmp)
            old = Workbook()
            old.active.title = "Гант"
            old.save(incoming / "home-internet.xlsx")
            unified = incoming / "plan.xlsx"
            mod.write_workbook(
                unified,
                entries=[
                    {
                        "team": "Тарифы",
                        "task": "Карточка",
                        "resource": "Косенко Данил",
                        "role": "FE",
                        "ticket": "",
                        "type": "деливери",
                        "weeks": {"2026-08-10": "3"},
                    }
                ],
                include_examples=False,
            )
            parsed = mod.parse_incoming(incoming)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["team"], "Тарифы")
        self.assertEqual(parsed[0]["resource"], "Косенко Данил")

    def test_vacation_column_keeps_real_role(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "План"
        ws.append(mod.PLAN_HEADERS)
        ws.append(
            [
                "Монетизация",
                "Отпуск",
                "",
                "",
                "Роган Татьяна",
                "SA",
                "да",
                date(2026, 9, 7),
                date(2026, 9, 11),
                "",
                "",
            ]
        )
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "vacation.xlsx"
            wb.save(path)
            parsed = mod.parse_workbook(path)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["role"], "SA")
        self.assertTrue(parsed[0]["vacation"])
        self.assertEqual(parsed[0]["weeks"]["2026-09-07"], "отпуск")
        self.assertEqual(parsed[0]["type"], "")

    def test_export_strips_cko_and_splits_vacation_from_role(self):
        entries = [
            {
                "team": "Тарифы",
                "task": "Карточка",
                "resource": "Косенко Данил",
                "role": "FE",
                "ticket": "",
                "type": "ЦКО",
                "weeks": {"2026-08-10": "3"},
            },
            {
                "team": "Монетизация",
                "task": "Отпуск",
                "resource": "Роган Татьяна",
                "role": "Отпуск",
                "ticket": "",
                "type": "",
                "weeks": {"2026-09-07": "отпуск"},
            },
        ]
        exported = []
        for item in entries:
            exported.extend(mod.entries_for_template([item]))
        by_task = {item["task"]: item for item in exported}
        self.assertEqual(by_task["Карточка"]["type"], "")
        self.assertFalse(by_task["Карточка"]["vacation"])
        vac = by_task["Отпуск"]
        self.assertTrue(vac["vacation"])
        self.assertEqual(vac["role"], "SA")


if __name__ == "__main__":
    unittest.main()
