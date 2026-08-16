#!/usr/bin/env python3
"""Unified Excel calendarization template and parser.

Subordinates fill the План sheet. Occupancy is dates and numbers, never cell color.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
ROSTER_PATH = ROOT / "roster.json"
DATA_PATH = ROOT / "data-h2-2026.json"
TEMPLATE_DIR = ROOT / "templates"

PLAN_YEAR = 2026
PLAN_START = date(2026, 6, 29)
PLAN_END = date(2026, 12, 28)

TEAMS = [
    {"key": "Монетизация", "full": "Монетизация", "color": "2E75B6", "bg": "DDEBF7"},
    {"key": "ДИ", "full": "Домашний интернет", "color": "C65911", "bg": "FCE4D6"},
    {"key": "ДГП", "full": "ДГП", "color": "BF9000", "bg": "FFF2CC"},
    {"key": "МегаИнтернет", "full": "МегаИнтернет", "color": "548235", "bg": "E2EFDA"},
    {"key": "Тарифы", "full": "Тарифы", "color": "C00000", "bg": "F8CBAD"},
]

TEAM_ALIASES = {
    "монетизация": "Монетизация",
    "репрайсы": "Монетизация",
    "домашний интернет": "ДИ",
    "ди": "ДИ",
    "фикса": "ДИ",
    "дгп": "ДГП",
    "мегаинтернет": "МегаИнтернет",
    "тарифы": "Тарифы",
    "тарифы (цко)": "Тарифы",
    "цко": "Тарифы",
}

ROLES = ["Дизайн", "SA", "PO", "BE", "FE", "QA", "Контент"]
ROLE_ALIASES = {
    "дизайн": "Дизайн",
    "дизайнер": "Дизайн",
    "dux": "Дизайн",
    "диз": "Дизайн",
    "sa": "SA",
    "аналитика": "SA",
    "аналитик": "SA",
    "po": "PO",
    "be": "BE",
    "бэк": "BE",
    "back": "BE",
    "ве": "BE",
    "fe": "FE",
    "фронт": "FE",
    "front": "FE",
    "qa": "QA",
    "тест": "QA",
    "тесты": "QA",
    "test": "QA",
    "контент": "Контент",
    "копирайт": "Контент",
}
TYPES = ["деливери", "дискавери"]
NEED_RESOURCES = [
    "Потребность дизайн",
    "Потребность SA",
    "Потребность BE",
    "Потребность FE",
    "Потребность QA",
    "Потребность контент",
]
NEED_ROLES = {
    "Потребность дизайн": "Дизайн",
    "Потребность SA": "SA",
    "Потребность BE": "BE",
    "Потребность FE": "FE",
    "Потребность QA": "QA",
    "Потребность контент": "Контент",
}
NEED_ALIASES = [
    (r"потребность\s*(дизайн|dux|диз)", "Потребность дизайн"),
    (r"потребность\s*(sa|аналитик)", "Потребность SA"),
    (r"потребность\s*(be|back|бэк|ве)", "Потребность BE"),
    (r"потребность\s*(fe|front|фронт)", "Потребность FE"),
    (r"потребность\s*(qa|test|тест)", "Потребность QA"),
    (r"потребность\s*контент", "Потребность контент"),
]
RESOURCE_CANONICAL = [
    (r"шлот?гауэр", "Шлотгауэр Иван"),
    (r"касенко|косенко", "Косенко Данил"),
    (r"савлук", "Савлук Богдан"),
    (r"успенский", "Успенский Павел"),
    (r"мерзликин", "Мерзликин Антон"),
    (r"роган|таня роган", "Роган Татьяна"),
    (r"судариков", "Судариков Алексей"),
    (r"колотыгин", "Колотыгин Никита"),
    (r"жогина", "Жогина Екатерина"),
    (r"папенко", "Папенко Руслан"),
    (r"гоголюк|\bгоша\b", "Гоголюк Георгий"),
    (r"титов", "Титов Иван"),
    (r"полубояринов", "Полубояринов Владимир"),
    (r"кувшинов", "Кувшинов Михаил"),
    (r"крюков", "Крюков Андрей"),
    (r"ворфоломеев", "Ворфоломеев Кирилл"),
    (r"ф[её]дорова", "Фёдорова Ольга"),
    (r"бодров", "Бодров Михаил"),
    (r"аксенов", "Аксенов Владимир"),
    (r"викулин", "Викулин Виталий"),
    (r"дмитриева|\bмаша\b", "Дмитриева Мария"),
    (r"фомин", "Фомин Антон"),
    (r"комаров", "Комаров Илья"),
    (r"калинкин", "Калинкин"),
    (r"юрасов", "Юрасов"),
    (r"крымов", "Крымов"),
]
EXTRA_PEOPLE = [
    ("Калинкин", "Дизайн"),
    ("Юрасов", "Дизайн"),
    ("Крымов", "Контент"),
]

META_HEADERS = [
    "Команда",
    "Задача",
    "Тикет",
    "Тип",
    "Исполнитель",
    "Роль",
    "Отпуск",
    "Начало",
    "Окончание",
    "Чел-дни",
    "Примечание",
]
PLAN_HEADERS = list(META_HEADERS)
VACATION_YES = {"да", "yes", "true", "1", "x", "+", "отпуск", "✓"}

HEADER_FILL = PatternFill("solid", fgColor="1C2531")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())


def iso(d: date) -> str:
    return d.isoformat()


def plan_weeks() -> list[date]:
    weeks = []
    day = monday_of(PLAN_START)
    last = monday_of(PLAN_END)
    while day <= last:
        weeks.append(day)
        day += timedelta(days=7)
    return weeks


def week_label(d: date) -> str:
    return f"{d.day:02d}.{d.month:02d}"


def canonical_team(raw: str) -> str:
    text = re.sub(r"\s+", " ", (raw or "").strip())
    if not text:
        return ""
    return TEAM_ALIASES.get(text.lower().replace("ё", "е"), text)


def canonical_role(raw: str) -> str:
    text = re.sub(r"\s+", " ", (raw or "").strip())
    if not text:
        return ""
    low = text.lower().replace("ё", "е")
    if low == "отпуск":
        return ""
    if text in ROLES:
        return text
    return ROLE_ALIASES.get(low, text)


def canonical_type(raw: str) -> str:
    text = re.sub(r"\s+", " ", (raw or "").strip())
    if not text:
        return ""
    low = text.lower().replace("ё", "е")
    if low in {"цко", "cko"}:
        return ""
    if text in TYPES:
        return text
    if low in {"деливери", "delivery"}:
        return "деливери"
    if low in {"дискавери", "discovery"}:
        return "дискавери"
    return ""


def parse_vacation_flag(value) -> bool:
    if value is True:
        return True
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value) == 1
    text = str(value or "").strip().lower().replace("ё", "е")
    return text in VACATION_YES


def short_official_name(name: str) -> str:
    text = re.sub(r"\s+", " ", (name or "").strip())
    parts = text.split()
    if len(parts) >= 3 and re.search(r"(вич|вна|ична|оглы|кызы)$", parts[-1], re.I):
        return f"{parts[0]} {parts[1]}"
    return text


def canonical_resource(resource: str) -> str:
    text = re.sub(r"\s+", " ", (resource or "").strip())
    if not text:
        return text
    low = text.lower().replace("ё", "е")
    for pattern, name in NEED_ALIASES:
        if re.search(pattern, low):
            return name
    for pattern, name in RESOURCE_CANONICAL:
        if re.search(pattern, low):
            return name
    return short_official_name(text)


def occupancy_cell(raw) -> str:
    if raw is None or raw == "":
        return ""
    if isinstance(raw, bool):
        return "1" if raw else ""
    if isinstance(raw, datetime):
        return ""
    if isinstance(raw, (int, float)):
        n = int(raw)
        return str(n) if n > 0 else ""
    text = str(raw).strip()
    if not text:
        return ""
    if "отпуск" in text.lower():
        return "отпуск"
    match = re.match(r"^(\d+)", text)
    if match:
        n = int(match.group(1))
        return str(n) if n > 0 else ""
    return "1"


def add_day(bucket: dict, day: date, value: str):
    if day.weekday() >= 5:
        return
    week = iso(monday_of(day))
    slot = bucket.get(week)
    if not isinstance(slot, dict):
        slot = {}
        bucket[week] = slot
    slot[iso(day)] = value or ""


def finalize_weeks(bucket: dict) -> dict:
    out = {}
    for week, payload in bucket.items():
        values = list(payload.values()) if isinstance(payload, dict) else list(payload or [])
        if not values:
            continue
        work = sum(1 for value in values if "отпуск" not in (value or "").lower())
        out[week] = str(work) if work else "отпуск"
    return out


def weeks_from_range(start: date, end: date, person_days: int | None = None, vacation: bool = False) -> dict:
    if end < start:
        start, end = end, start
    days = []
    cursor = start
    while cursor <= end:
        if cursor.weekday() < 5:
            days.append(cursor)
        cursor += timedelta(days=1)
    if person_days is not None:
        days = days[: max(0, int(person_days))]
    bucket = {}
    label = "отпуск" if vacation else "1"
    for day in days:
        add_day(bucket, day, label)
    return finalize_weeks(bucket)


def parse_date_value(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value)).date()
        except ValueError:
            return None
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def load_roster_people() -> list[tuple[str, str, str]]:
    people = []
    seen = set()
    if ROSTER_PATH.exists():
        roster = json.loads(ROSTER_PATH.read_text(encoding="utf-8"))
        for person in roster.get("people", []):
            name = canonical_resource(person.get("name") or "")
            role = canonical_role(person.get("role") or "")
            teams = ", ".join(
                next((team["full"] for team in TEAMS if team["key"] == alloc.get("team")), alloc.get("team", ""))
                for alloc in person.get("allocations") or []
            )
            if name and name not in seen:
                seen.add(name)
                people.append((name, role, teams))
    for name, role in EXTRA_PEOPLE:
        name = canonical_resource(name)
        if name not in seen:
            seen.add(name)
            people.append((name, role, ""))
    for name in NEED_RESOURCES:
        people.append((name, NEED_ROLES[name], "потребность"))
    return people


def role_for_resource(resource: str, fallback: str = "") -> str:
    people = load_roster_people()
    resource = canonical_resource(resource)
    low = (resource or "").lower().replace("ё", "е")
    for name, role, _teams in people:
        if name.lower().replace("ё", "е") == low:
            return role
    mapped = canonical_role(fallback)
    if mapped in ROLES:
        return mapped
    return mapped or "Другое"


def display_team(key: str) -> str:
    for team in TEAMS:
        if team["key"] == key:
            return team["full"]
    return key


def person_days_of(weeks: dict) -> int:
    total = 0
    for value in (weeks or {}).values():
        cell = occupancy_cell(value)
        if cell and cell != "отпуск":
            total += int(cell)
    return total


def first_last_dates(weeks: dict) -> tuple[date | None, date | None]:
    isos = sorted((weeks or {}).keys())
    if not isos:
        return None, None
    start = date.fromisoformat(isos[0])
    last_monday = date.fromisoformat(isos[-1])
    last_value = occupancy_cell(weeks[isos[-1]])
    if last_value == "отпуск":
        end = last_monday + timedelta(days=4)
    else:
        days = int(last_value) if last_value.isdigit() else 5
        end = last_monday + timedelta(days=min(4, max(0, days - 1)))
    return start, end


def entries_for_template(entries: list[dict]) -> list[dict]:
    out = []
    for entry in entries or []:
        role_raw = str(entry.get("role") or "").strip()
        role_is_vacation = role_raw.lower().replace("ё", "е") == "отпуск"
        role = role_for_resource(entry.get("resource") or "") if role_is_vacation else canonical_role(role_raw)
        typ = canonical_type(entry.get("type") or "")
        weeks = entry.get("weeks") or {}
        work = {
            week: occupancy_cell(value)
            for week, value in weeks.items()
            if occupancy_cell(value) and occupancy_cell(value) != "отпуск"
        }
        vac = {week: "отпуск" for week, value in weeks.items() if occupancy_cell(value) == "отпуск"}
        if role_is_vacation:
            vac = {week: "отпуск" for week, value in weeks.items() if occupancy_cell(value)}
            work = {}
        base = {
            **entry,
            "role": role,
            "type": typ,
        }
        if work:
            out.append({**base, "weeks": work, "vacation": False})
        if vac:
            out.append({**base, "weeks": vac, "vacation": True, "role": role or role_for_resource(entry.get("resource") or "")})
        if not work and not vac and entry.get("vacation"):
            out.append({**base, "vacation": True})
    return out


def write_workbook(path: Path, entries: list[dict] | None = None, include_examples: bool = True) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    weeks = plan_weeks()
    headers = META_HEADERS + [week_label(week) for week in weeks]
    people = load_roster_people()

    wb = Workbook()

    ws_help = wb.active
    ws_help.title = "Инструкция"
    _write_instructions(ws_help)

    ws_ref = wb.create_sheet("Справочник")
    _write_reference(ws_ref, people)

    ws = wb.create_sheet("План")
    ws.append(headers)
    for col, _title in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN

    rows_out = []
    if include_examples:
        rows_out.extend(_example_rows(weeks))
    for entry in entries_for_template(entries or []):
        rows_out.append(_entry_to_row(entry, weeks))

    empty_target = max(120, len(rows_out) + 40)
    while len(rows_out) < empty_target:
        rows_out.append([""] * len(headers))

    for values in rows_out:
        ws.append(values)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = THIN
            cell.font = Font(name="Calibri", size=11)
            if cell.column in (8, 9):
                cell.number_format = "DD.MM.YYYY"
            if cell.column >= 12:
                cell.alignment = Alignment(horizontal="center")
        task = str(ws.cell(row[0].row, 2).value or "")
        if task.upper().startswith("ПРИМЕР"):
            for cell in row:
                cell.fill = EXAMPLE_FILL

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.row_dimensions[1].height = 32
    widths = {
        1: 22,
        2: 42,
        3: 16,
        4: 14,
        5: 24,
        6: 12,
        7: 10,
        8: 13,
        9: 13,
        10: 10,
        11: 18,
    }
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 7)

    last_col = get_column_letter(len(headers))
    last_row = ws.max_row
    people_end = max(2, len(people) + 1)
    roles_end = 1 + len(ROLES)

    dv_team = DataValidation(type="list", formula1="=Справочник!$A$2:$A$6", allow_blank=True)
    dv_type = DataValidation(type="list", formula1="=Справочник!$C$2:$C$3", allow_blank=True)
    dv_person = DataValidation(type="list", formula1=f"=Справочник!$E$2:$E${people_end}", allow_blank=True)
    dv_role = DataValidation(type="list", formula1=f"=Справочник!$G$2:$G${roles_end}", allow_blank=True)
    dv_vacation = DataValidation(type="list", formula1='"да"', allow_blank=True)
    dv_week = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    for dv in (dv_team, dv_type, dv_person, dv_role, dv_vacation, dv_week):
        dv.showErrorMessage = True
        dv.errorTitle = "Неверное значение"
        dv.error = "Выберите значение из списка — так не будет разночтений."
        ws.add_data_validation(dv)
    dv_team.add(f"A2:A{last_row}")
    dv_type.add(f"D2:D{last_row}")
    dv_person.add(f"E2:E{last_row}")
    dv_role.add(f"F2:F{last_row}")
    dv_vacation.add(f"G2:G{last_row}")
    dv_week.add(f"L2:{last_col}{last_row}")

    for team in TEAMS:
        fill = PatternFill("solid", fgColor=team["bg"])
        font = Font(color=team["color"], bold=True, name="Calibri")
        ws.conditional_formatting.add(
            f"A2:A{last_row}",
            FormulaRule(formula=[f'$A2="{team["full"]}"'], fill=fill, font=font),
        )

    for col, week in enumerate(weeks, 12):
        ws.cell(1, col).comment = None
        ws.cell(1, col).value = week_label(week)

    ws_help.sheet_properties.tabColor = "3DD6C3"
    ws.sheet_properties.tabColor = "2E75B6"
    ws_ref.sheet_state = "visible"
    ws_ref.protection.sheet = True

    wb.save(path)
    return path


def _write_instructions(ws):
    ws.sheet_view.showGridLines = False
    title = ws["A1"]
    title.value = "Единый формат календаризации 3–4Q 2026"
    title.font = Font(name="Calibri", size=18, bold=True, color="1C2531")
    ws.merge_cells("A1:G1")
    lines = [
        "",
        "Зачем этот файл",
        "Все команды заполняют один и тот же Excel. Цвета ячеек больше не означают занятость. Занятость — это даты и числа дней. Отпуск отмечается отдельно.",
        "",
        "Что заполнять",
        "Только лист «План». Жёлтые строки — полный пример из таймплана МегаИнтернета. Удалите их перед отправкой, в таймплан они не попадут.",
        "",
        "Колонки",
        "Команда — только из списка: Монетизация, Домашний интернет, ДГП, МегаИнтернет, Тарифы. Фикса = Домашний интернет, Репрайсы = Монетизация.",
        "Задача — одно название на одну строку исполнителя. Несколько людей на задачу = несколько строк.",
        "Тикет — B2CPROD-… / CKO-…, если есть.",
        "Тип — только деливери или дискавери, либо пусто. ЦКО не пишите.",
        "Исполнитель — только из списка. Не пишите FE, бэк, DUX, Касенко, Шлогауэр.",
        "Роль — Дизайн / SA / PO / BE / FE / QA / Контент. Если не указать, подставится роль из штата. Отпуск ролью не является.",
        "Отпуск — поставьте «да», если эта строка про нерабочие дни человека. Роль оставьте настоящей (QA, SA, …). Даты начала и окончания — период отпуска.",
        "Начало и Окончание — рабочие даты включительно. Выходные игнорируются.",
        "Чел-дни — необязательно и только для работы, не для отпуска. Если пусто, считаются все будни между началом и окончанием. Если указано 3 при пяти днях в диапазоне, в план попадут первые 3 будня.",
        "Недели справа — занятость работой: 1–5. Если заполнены, они важнее дат. В отпускные недели ничего не ставьте: отпуск уже отмечен колонкой «Отпуск».",
        "",
        "Чего нельзя делать",
        "Не закрашивайте недели цветом вместо числа. Цвет команды в колонке «Команда» ставится сам и в расчёт не идёт.",
        "Не пишите свободным текстом роли (аналитика, фронт, тесты) и не сокращайте имена.",
        "Не ставьте одну строку на всю команду. Одна строка = один исполнитель × одна задача.",
        "Потребность без человека — выберите «Потребность SA/FE/BE/QA/дизайн/контент». Это спрос без FTE.",
        "",
        "Отпуска",
        "Отдельная строка на человека: роль из списка, в колонке «Отпуск» — да, даты периода. Эти дни у человека не рабочие и в спрос не идут.",
        "",
        "Как отдать файл",
        "Сохраните как .xlsx и пришлите целиком. Можно один общий файл на все команды или по файлу на команду — структура листа «План» должна совпадать.",
    ]
    for idx, line in enumerate(lines, 2):
        cell = ws.cell(idx, 1, line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line in {"Зачем этот файл", "Что заполнять", "Колонки", "Чего нельзя делать", "Отпуска", "Как отдать файл"}:
            cell.font = Font(name="Calibri", size=13, bold=True, color="1C2531")
        else:
            cell.font = Font(name="Calibri", size=11, color="1C2531")
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=7)
        ws.row_dimensions[idx].height = 36 if len(line) > 80 else 20
    ws.column_dimensions["A"].width = 28
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 16


def _write_reference(ws, people):
    ws["A1"] = "Команда"
    ws["C1"] = "Тип"
    ws["E1"] = "Исполнитель"
    ws["F1"] = "Роль штата"
    ws["G1"] = "Роль"
    ws["H1"] = "Команды исполнителя"
    ws["I1"] = "Отпуск"
    for cell in ("A1", "C1", "E1", "F1", "G1", "H1", "I1"):
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    for idx, team in enumerate(TEAMS, 2):
        ws.cell(idx, 1, team["full"])
        ws.cell(idx, 1).fill = PatternFill("solid", fgColor=team["bg"])
        ws.cell(idx, 1).font = Font(color=team["color"], bold=True)
    for idx, typ in enumerate(TYPES, 2):
        ws.cell(idx, 3, typ)
    for idx, role in enumerate(ROLES, 2):
        ws.cell(idx, 7, role)
    ws.cell(2, 9, "да")
    for idx, (name, role, teams) in enumerate(people, 2):
        ws.cell(idx, 5, name)
        ws.cell(idx, 6, role)
        ws.cell(idx, 8, teams)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 28
    ws.column_dimensions["I"].width = 12
    note_row = max(22, len(people) + 4)
    note = ws.cell(note_row, 1, "Цвета команд совпадают с дашбордом. Они справочные: занятость цветом не кодируется.")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)


def _megainternet_example_entries() -> list[dict]:
    if DATA_PATH.exists():
        payload = json.loads(DATA_PATH.read_text(encoding="utf-8"))
        rows = []
        for item in payload.get("entries") or []:
            if item.get("team") != "МегаИнтернет":
                continue
            task = re.sub(r"\s+", " ", str(item.get("task") or "").strip())
            if not task:
                continue
            if not task.upper().startswith("ПРИМЕР"):
                task = f"ПРИМЕР: {task}"
            rows.append({**item, "task": task, "team": "МегаИнтернет"})
        if rows:
            return rows
    return [
        {
            "team": "МегаИнтернет",
            "task": "ПРИМЕР: 5G кино",
            "resource": "Лера",
            "role": "Дизайн",
            "ticket": "",
            "type": "",
            "weeks": {"2026-08-03": "2", "2026-08-10": "5"},
        }
    ]


def _example_rows(weeks: list[date]) -> list[list]:
    week_index = {iso(week): idx for idx, week in enumerate(weeks)}
    return [_entry_to_row(item, weeks, week_index) for item in entries_for_template(_megainternet_example_entries())]


def _entry_to_row(entry: dict, weeks: list[date], week_index: dict | None = None) -> list:
    week_index = week_index or {iso(week): idx for idx, week in enumerate(weeks)}
    vacation = bool(entry.get("vacation"))
    occupancy = entry.get("weeks") or {}
    if vacation:
        occupancy = {week: "отпуск" for week in occupancy}
    start, end = first_last_dates(occupancy)
    row = [""] * (len(META_HEADERS) + len(weeks))
    row[0] = display_team(canonical_team(entry.get("team") or ""))
    row[1] = entry.get("task") or ""
    row[2] = entry.get("ticket") or ""
    row[3] = canonical_type(entry.get("type") or "")
    row[4] = canonical_resource(entry.get("resource") or "")
    row[5] = canonical_role(entry.get("role") or "") or role_for_resource(entry.get("resource") or "")
    row[6] = "да" if vacation else ""
    row[7] = start
    row[8] = end
    row[9] = None if vacation else (person_days_of(occupancy) or None)
    row[10] = ""
    if not vacation:
        for week_iso, value in occupancy.items():
            idx = week_index.get(week_iso)
            if idx is None:
                continue
            cell = occupancy_cell(value)
            if cell.isdigit():
                row[len(META_HEADERS) + idx] = int(cell)
    return row


def parse_workbook(path: Path) -> list[dict]:
    wb = load_workbook(Path(path), data_only=False)
    if "План" not in wb.sheetnames:
        raise ValueError(f"{path}: нет листа «План»")
    ws = wb["План"]
    header = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(header)}
    week_cols = {}
    for idx, title in enumerate(header):
        if re.fullmatch(r"\d{2}\.\d{2}", title):
            day, month = title.split(".")
            week_cols[idx] = iso(monday_of(date(PLAN_YEAR, int(month), int(day))))
        elif re.fullmatch(r"\d{4}-\d{2}-\d{2}", title):
            week_cols[idx] = iso(monday_of(date.fromisoformat(title)))

    def take(values, name, default=""):
        idx = col.get(name)
        if idx is None or idx >= len(values):
            return default
        return values[idx]

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row) + [None] * max(0, len(header) - len(row))
        team = canonical_team(take(values, "Команда"))
        task = re.sub(r"\s+", " ", str(take(values, "Задача") or "").strip())
        if not team or not task:
            continue
        if task.upper().startswith("ПРИМЕР"):
            continue
        ticket = str(take(values, "Тикет") or "").strip()
        typ = canonical_type(take(values, "Тип"))
        resource = canonical_resource(str(take(values, "Исполнитель") or "").strip())
        role_raw = str(take(values, "Роль") or "").strip()
        role = canonical_role(role_raw)
        vacation = parse_vacation_flag(take(values, "Отпуск")) or role_raw.lower().replace("ё", "е") == "отпуск"
        start = parse_date_value(take(values, "Начало", None))
        end = parse_date_value(take(values, "Окончание", None))
        person_days_raw = take(values, "Чел-дни", None)
        person_days = None
        if isinstance(person_days_raw, (int, float)) and not isinstance(person_days_raw, bool):
            person_days = int(person_days_raw)
        elif str(person_days_raw or "").strip().isdigit():
            person_days = int(str(person_days_raw).strip())
        note = str(take(values, "Примечание") or "").strip()
        if "отпуск" in note.lower():
            vacation = True

        week_map = {}
        for col_idx, week_iso in week_cols.items():
            if col_idx >= len(values):
                continue
            cell = occupancy_cell(values[col_idx])
            if cell:
                week_map[week_iso] = cell

        if vacation and week_map:
            week_map = {week: "отпуск" for week in week_map}
        if not week_map:
            if not start or not end:
                continue
            week_map = weeks_from_range(start, end, person_days=None if vacation else person_days, vacation=vacation)
        elif vacation:
            week_map = {week: "отпуск" for week in week_map}
        if not week_map:
            continue
        if not role:
            role = role_for_resource(resource)
        if not resource:
            resource = role or "—"
        entries.append(
            {
                "team": team,
                "task": task,
                "resource": resource,
                "role": role or "Другое",
                "ticket": ticket,
                "type": typ,
                "vacation": vacation,
                "weeks": week_map,
            }
        )
    return entries


def is_unified_workbook(path: Path) -> bool:
    try:
        wb = load_workbook(Path(path), read_only=True)
        names = wb.sheetnames
        wb.close()
        return "План" in names
    except Exception:
        return False


def parse_incoming(incoming: Path) -> list[dict]:
    files = sorted(Path(incoming).glob("*.xlsx"))
    plans = [path for path in files if is_unified_workbook(path)]
    entries = []
    for path in plans:
        entries.extend(parse_workbook(path))
    return entries


def write_default_templates(data_path: Path | None = None) -> tuple[Path, Path]:
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    empty = TEMPLATE_DIR / "Календаризация_шаблон.xlsx"
    filled = TEMPLATE_DIR / "Календаризация_текущий_план.xlsx"
    write_workbook(empty, entries=None, include_examples=True)
    current = []
    source = Path(data_path or DATA_PATH)
    if source.exists():
        payload = json.loads(source.read_text(encoding="utf-8"))
        current = payload.get("entries") or []
    write_workbook(filled, entries=current, include_examples=False)
    shutil.copy2(empty, TEMPLATE_DIR / "calendarization_template.xlsx")
    shutil.copy2(filled, TEMPLATE_DIR / "calendarization_current_plan.xlsx")
    return empty, filled


def main():
    parser = argparse.ArgumentParser(description="Собрать или разобрать единый Excel календаризации")
    parser.add_argument("--from-data", action="store_true", help="заполнить текущим data-h2-2026.json")
    parser.add_argument("--parse", type=Path, help="прочитать заполненный файл и напечатать число строк")
    args = parser.parse_args()
    if args.parse:
        rows = parse_workbook(args.parse)
        print(f"{args.parse}: {len(rows)} строк")
        return
    empty, filled = write_default_templates()
    print(f"wrote {empty}")
    print(f"wrote {filled}")


if __name__ == "__main__":
    main()
