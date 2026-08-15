#!/usr/bin/env python3
"""Build a unified weekly calendarization JSON from incoming team plans."""
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
INCOMING = ROOT / "incoming"
OUT = ROOT / "data-h2-2026.json"

MONTHS_RU = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5, "июнь": 6,
    "июль": 7, "август": 8, "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}

ROLE_ALIASES = [
    (r"dux|дизайн|дизайнер|\(диз\)|диз\b", "Дизайн"),
    (r"\bsa\b|аналитик|аналитика|adr", "SA"),
    (r"\bbe\b|бэк|back|\bве\b|\(ве\)", "BE"),
    (r"\bfe\b|фронт|front", "FE"),
    (r"qa\s*\(s\)", "QA"),
    (r"\bqa\b|тест|test", "QA"),
    (r"копирайт", "Копирайт"),
    (r"контент", "Контент"),
    (r"\bux\b", "UX"),
    (r"\brnd\b", "RnD"),
    (r"менеджер", "Менеджер"),
    (r"ресерч|ресёрч|research", "SA"),
]


def monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())


def iso(d: date) -> str:
    return d.isoformat()


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def cell_value(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    text = re.sub(r"\s+", " ", str(v)).strip()
    return "" if text in {"None", "-", "•"} else text


def fill_theme(cell) -> tuple[str | None, int | None, float]:
    fill = cell.fill
    if not fill or fill.patternType in (None, "none"):
        return None, None, 0.0
    fg = fill.fgColor
    if fg is None:
        return None, None, 0.0
    tint = float(fg.tint or 0)
    return fg.type, getattr(fg, "theme", None), tint


def has_fill(cell) -> bool:
    fill = cell.fill
    if not fill or fill.patternType in (None, "none"):
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    if fg.type == "rgb":
        rgb = (fg.rgb or "")[-6:].upper()
        return rgb not in {"", "000000", "FFFFFF", "00000000"}
    if fg.type == "theme":
        # Theme 2 is the weekend/grid gray used in DGP, not occupancy.
        if fg.theme == 2:
            return False
        if fg.theme in (0, 1) and abs(float(fg.tint or 0)) < 0.01:
            return False
        return True
    if fg.type == "indexed":
        return fg.indexed not in (64, 65, 0, 1)
    return True


def occupied(cell) -> bool:
    return bool(cell_value(cell)) or has_fill(cell)


def normalize_role(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    low = text.lower().replace("ё", "е")
    for pattern, role in ROLE_ALIASES:
        if re.search(pattern, low, re.I):
            if "qa (s)" in low or "qa(s)" in low:
                return "QA"
            return role
    return text


def parse_role_and_name(label: str) -> tuple[str, str]:
    text = re.sub(r"\s+", " ", (label or "").strip())
    if not text:
        return "", ""
    low = text.lower()
    prefixes = [
        ("дизайнер ", "Дизайн"),
        ("дизайн ", "Дизайн"),
        ("d ", "Дизайн"),
        ("sa ", "SA"),
        ("be ", "BE"),
        ("fe ", "FE"),
        ("qa ", "QA"),
    ]
    for prefix, role in prefixes:
        if low.startswith(prefix):
            return role, text[len(prefix):].strip() or text
    role = normalize_role(text)
    if role in KNOWN_ROLES and (text.startswith("(") or " " not in text):
        return role, text.strip("() ") or role
    return role, text


def extract_ticket(text: str) -> str:
    match = re.search(r"(B2CPROD-\d+|CKO-\d+|LK-\d+|WLK-\d+)", text or "", re.I)
    return match.group(1).upper().replace("B2CPROD", "B2CPROD") if match else ""


def add_day(bucket: dict, day: date, value: str):
    week = iso(monday_of(day))
    text = value.strip() if value else ""
    prev = bucket.get(week, [])
    if text and text not in prev:
        prev.append(text)
    elif not prev:
        prev.append("")
    bucket[week] = prev


def finalize_weeks(bucket: dict) -> dict:
    out = {}
    for week, parts in bucket.items():
        labels = [p for p in parts if p and p != "•"]
        if not labels:
            out[week] = "•"
            continue
        if all("отпуск" in p.lower() for p in labels):
            out[week] = "отпуск"
        else:
            out[week] = " / ".join(labels[:3])[:80]
    return out


Q3_START = date(2026, 6, 29)
Q3_END = date(2026, 12, 28)


def clip_weeks(weeks: dict, start: date = Q3_START, end: date = Q3_END) -> dict:
    start_iso, end_iso = iso(start), iso(end)
    return {k: v for k, v in weeks.items() if start_iso <= k <= end_iso}


def entry(team, task, resource, role, ticket="", typ="", weeks=None):
    weeks = {k: v for k, v in (weeks or {}).items() if v}
    if not weeks:
        return None
    task = re.sub(r"\s+", " ", (task or "").strip()) or "Без названия"
    resource = re.sub(r"\s+", " ", (resource or "").strip()) or role or "—"
    if resource.startswith("(") and resource.endswith(")"):
        inner = resource[1:-1].strip()
        if normalize_role(inner) == (role or normalize_role(inner)):
            resource = role or inner
    return {
        "team": team,
        "task": task,
        "resource": resource,
        "role": role or "Другое",
        "ticket": ticket or extract_ticket(task),
        "type": typ or "",
        "weeks": weeks,
    }


def merge_entries(entries: list[dict]) -> list[dict]:
    grouped: dict[tuple, dict] = {}
    order: list[tuple] = []
    for item in entries:
        key = (item["team"], item["task"], item["resource"], item["role"], item["ticket"], item["type"])
        if key not in grouped:
            grouped[key] = {**item, "weeks": dict(item["weeks"])}
            order.append(key)
            continue
        weeks = grouped[key]["weeks"]
        for week, value in item["weeks"].items():
            prev = weeks.get(week)
            if not prev or prev == "•":
                weeks[week] = value
    return [grouped[key] for key in order]


SKIP_RESOURCES = {"менеджер", "катя (менеджер)", "менеджер "}
KNOWN_ROLES = {"Дизайн", "SA", "BE", "FE", "QA", "Копирайт", "Контент", "Менеджер", "Отпуск", "UX", "RnD"}


def dgp_day_label(cell, day: date) -> str | None:
    val = cell_value(cell)
    if val:
        return val
    if day.weekday() >= 5 or day.day == 1:
        return None
    if not has_fill(cell):
        return None
    kind, theme, tint = fill_theme(cell)
    if kind == "theme" and theme == 0 and tint < 0:
        return "отпуск"
    return "•"


def parse_monetization() -> list[dict]:
    wb = load_workbook(INCOMING / "monetization.xlsx", data_only=False)
    ws = wb.active
    month_at = {}
    current = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if isinstance(v, str) and v.strip().lower() in MONTHS_RU:
            current = MONTHS_RU[v.strip().lower()]
        if current:
            month_at[col] = current
    col_dates = {}
    for col in range(4, ws.max_column + 1):
        day = ws.cell(2, col).value
        month = month_at.get(col)
        if month and isinstance(day, (int, float)):
            try:
                col_dates[col] = date(2026, month, int(day))
            except ValueError:
                continue

    rows = []
    current_type = ""
    current_task = ""
    current_ticket = ""
    for r in range(3, ws.max_row + 1):
        a = cell_value(ws.cell(r, 1))
        b = cell_value(ws.cell(r, 2))
        c = cell_value(ws.cell(r, 3))
        if a:
            current_type = a
        if a in {"деливери", "дискавери"} and b:
            current_task = b
            current_ticket = c
            continue
        if a == "отпуска" or (not a and b and current_type == "отпуска"):
            name = b or a
            bucket = defaultdict(list)
            for col, day in col_dates.items():
                if occupied(ws.cell(r, col)):
                    add_day(bucket, day, "отпуск")
            item = entry("Монетизация", "Отпуск", name, "Отпуск", weeks=clip_weeks(finalize_weeks(bucket), Q3_START))
            if item:
                rows.append(item)
            continue
        role_src = b.lower() if b else ""
        if role_src in {"дизайн", "аналитика", "бэк", "фронт", "тесты", "копирайт"}:
            role = normalize_role(b)
            bucket = defaultdict(list)
            for col, day in col_dates.items():
                cell = ws.cell(r, col)
                if occupied(cell):
                    add_day(bucket, day, cell_value(cell) or b)
            item = entry(
                "Монетизация",
                current_task,
                b,
                role,
                current_ticket or c,
                current_type,
                clip_weeks(finalize_weeks(bucket), Q3_START),
            )
            if item:
                rows.append(item)
    return rows


def parse_megainternet() -> list[dict]:
    wb = load_workbook(INCOMING / "megainternet.xlsx", data_only=True)
    ws = wb["Таймплан"]
    month_at = {}
    current = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if isinstance(v, str) and v.strip().lower() in MONTHS_RU:
            current = MONTHS_RU[v.strip().lower()]
        if current and col >= 7:
            month_at[col] = current

    col_week = {}
    for col in range(7, ws.max_column + 1):
        label = cell_value(ws.cell(2, col))
        month = month_at.get(col)
        if not label or "-" not in label or not month:
            continue
        start_s, end_s = label.split("-", 1)
        try:
            start_d, end_d = int(start_s), int(end_s)
        except ValueError:
            continue
        year, m = 2026, month
        if start_d > 20 and end_d < 15:
            m = month - 1 if month > 1 else 12
            year = 2026 if month > 1 else 2025
        try:
            col_week[col] = iso(monday_of(date(year, m, start_d)))
        except ValueError:
            for guess in range(28, 32):
                try:
                    col_week[col] = iso(monday_of(date(year, m, guess)))
                    break
                except ValueError:
                    continue

    rows = []
    current_task = ""
    current_ticket = ""
    for r in range(4, ws.max_row + 1):
        a = cell_value(ws.cell(r, 1))
        b = cell_value(ws.cell(r, 2))
        if a:
            ticket_match = re.search(r"(B2CPROD-\d+|CKO-\d+|LK-\d+)", a, re.I)
            current_ticket = ticket_match.group(1) if ticket_match else ""
            current_task = re.sub(r"\s*\+?\s*https?://\S+", "", a).strip()
            continue
        if not b or b.lower().strip() in SKIP_RESOURCES:
            continue
        role, resource = parse_role_and_name(b)
        if resource.lower().strip() in SKIP_RESOURCES or role == "Менеджер":
            continue
        weeks = {}
        for col, week in col_week.items():
            val = cell_value(ws.cell(r, col))
            if not val:
                continue
            weeks[week] = val
        item = entry("МегаИнтернет", current_task, resource or b, role or normalize_role(b), current_ticket, "", clip_weeks(weeks, Q3_START))
        if item:
            rows.append(item)
    return rows


def parse_home_internet() -> list[dict]:
    wb = load_workbook(INCOMING / "home-internet.xlsx", data_only=True)
    ws = wb["Гант"]
    col_dates = {}
    for col in range(5, ws.max_column + 1):
        d = parse_date(ws.cell(3, col).value)
        if d:
            col_dates[col] = d

    grouped = {}
    current_task = ""
    current_ticket = ""
    for r in range(4, ws.max_row + 1):
        task = cell_value(ws.cell(r, 4))
        ticket = cell_value(ws.cell(r, 3))
        if task:
            current_task = task
            current_ticket = ticket
        for col, day in col_dates.items():
            val = cell_value(ws.cell(r, col))
            if not val or val.lower() == "релиз":
                continue
            role = normalize_role(val)
            if not role or role.lower() == "релиз":
                continue
            key = (current_task, role, val if val != role else role)
            grouped.setdefault(key, {"ticket": current_ticket, "days": defaultdict(list)})
            add_day(grouped[key]["days"], day, val)

    rows = []
    for (task, role, resource), payload in grouped.items():
        item = entry("ДИ", task, resource, role, payload["ticket"], "", clip_weeks(finalize_weeks(payload["days"]), Q3_START))
        if item:
            rows.append(item)
    return rows


TASK_HINT = re.compile(
    r"^(№|\[|q[234]|delivery|discovery|дгп|роли |гео |срок |скидки |любимые |мегасемья)",
    re.I,
)
ROLE_HEADERS = {"дизайн", "sa", "be", "fe", "qa", "q3", "q2 готово"}


def parse_dgp() -> list[dict]:
    wb = load_workbook(INCOMING / "dgp.xlsx", data_only=False)
    ws = wb["График 2026"]
    start = date(2026, 3, 16)
    col_dates = {col: start + timedelta(days=col - 2) for col in range(2, ws.max_column + 1)}

    rows = []
    current_task = ""
    current_role = ""
    for r in range(3, ws.max_row + 1):
        label = cell_value(ws.cell(r, 1))
        if not label:
            continue
        low = label.lower().strip()
        if low in {"женя"}:
            continue
        if low in ROLE_HEADERS:
            current_role = normalize_role(label)
            continue
        if TASK_HINT.search(low) or (len(label) > 28 and " " in label and not re.match(r"^(sa|be|fe|qa|d|дизайнер)\b", low)):
            current_task = label
            current_role = ""
            continue
        role, resource = parse_role_and_name(label)
        if not resource or resource.lower() in ROLE_HEADERS:
            if role in KNOWN_ROLES:
                current_role = role
            continue
        if current_role and role not in KNOWN_ROLES:
            role = current_role
        bucket = defaultdict(list)
        for col, day in col_dates.items():
            label = dgp_day_label(ws.cell(r, col), day)
            if label:
                add_day(bucket, day, label)
        item = entry(
            "ДГП",
            current_task or "ДГП",
            resource,
            role or current_role or "Другое",
            extract_ticket(current_task),
            "",
            clip_weeks(finalize_weeks(bucket), Q3_START),
        )
        if item:
            rows.append(item)
    return rows


def parse_tariffs_cko() -> list[dict]:
    import jpype
    import mpxj  # noqa: F401  # registers classpath

    if not jpype.isJVMStarted():
        jpype.startJVM()
    from org.mpxj.reader import UniversalProjectReader

    project = UniversalProjectReader().read(str(INCOMING / "tariffs-cko.mpp"))

    def parent_name(task):
        names = []
        cur = task
        for _ in range(10):
            cur = cur.getParentTask()
            if cur is None:
                break
            name = str(cur.getName() or "").strip()
            if name and name not in {"Орига v_2", "Орига", "ЦКО", "ЦКО Общий блок"}:
                names.append(name)
        for name in names:
            if not name.lower().startswith("цко"):
                return name
        return names[0] if names else "Тарифы ЦКО"

    def is_leaf(task):
        children = task.getChildTasks()
        return children is None or children.size() == 0

    def is_cko(task):
        name = str(task.getName() or "")
        if "ЦКО" in name:
            return True
        for assignment in task.getResourceAssignments():
            res = assignment.getResource()
            if res and "ЦКО" in str(res.getName() or ""):
                return True
        return False

    rows = []
    for task in project.getTasks():
        if not is_cko(task) or not is_leaf(task):
            continue
        name = str(task.getName() or "").strip()
        start = task.getStart()
        finish = task.getFinish()
        if start is None or finish is None:
            continue
        start_d = datetime(
            start.getYear(), start.getMonthValue(), start.getDayOfMonth()
        ).date() if hasattr(start, "getYear") else parse_date(start)
        finish_d = datetime(
            finish.getYear(), finish.getMonthValue(), finish.getDayOfMonth()
        ).date() if hasattr(finish, "getYear") else parse_date(finish)
        if not start_d or not finish_d:
            continue
        resources = []
        for assignment in task.getResourceAssignments():
            res = assignment.getResource()
            if res and res.getName():
                resources.append(str(res.getName()))
        resource = resources[0] if resources else "ЦКО"
        role = normalize_role(name) or normalize_role(resource) or "Другое"
        feature = parent_name(task)
        bucket = defaultdict(list)
        day = start_d
        while day <= finish_d:
            if day.weekday() < 5:
                add_day(bucket, day, role)
            day += timedelta(days=1)
        item = entry("Тарифы", feature, resource, role, "", "ЦКО", clip_weeks(finalize_weeks(bucket), Q3_START))
        if item:
            rows.append(item)
    return rows


TEAMS = {
    "Монетизация": {"full": "Монетизация", "color": "#2E75B6", "bg": "#DDEBF7"},
    "ДИ": {"full": "Домашний интернет", "color": "#C65911", "bg": "#FCE4D6"},
    "ДГП": {"full": "ДГП", "color": "#BF9000", "bg": "#FFF2CC"},
    "МегаИнтернет": {"full": "МегаИнтернет", "color": "#548235", "bg": "#E2EFDA"},
    "Тарифы": {"full": "Тарифы (ЦКО)", "color": "#C00000", "bg": "#F8CBAD"},
}


def main():
    entries = []
    sources = {
        "Монетизация": parse_monetization,
        "МегаИнтернет": parse_megainternet,
        "ДИ": parse_home_internet,
        "ДГП": parse_dgp,
        "Тарифы": parse_tariffs_cko,
    }
    for name, fn in sources.items():
        part = fn()
        print(f"{name}: {len(part)} entries")
        entries.extend(part)

    entries = merge_entries(entries)
    week_set = sorted({w for e in entries for w in e["weeks"]})
    weeks = [{"iso": w, "label": f"{w[8:10]}.{w[5:7]}"} for w in week_set]
    payload = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "title": "3–4Q 2026",
        "teams": TEAMS,
        "weeks": weeks,
        "entries": entries,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"wrote {OUT} weeks={len(weeks)} entries={len(entries)}")
    if weeks:
        print(f"span {weeks[0]['label']} — {weeks[-1]['label']}")


if __name__ == "__main__":
    main()
